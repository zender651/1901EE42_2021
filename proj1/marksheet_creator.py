import os
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, Alignment

current_directory = os.getcwd()
input_dir = os.path.join(
    current_directory, 'assets', 'input')
if not os.path.exists(input_dir):
    os.makedirs(input_dir)
ouptut_dir = os.path.join(
    current_directory, 'assets', 'output', 'marksheet')
if not os.path.exists(ouptut_dir):
    os.makedirs(ouptut_dir)
for f in os.listdir(input_dir):
    os.remove(os.path.join(input_dir, f))
for f in os.listdir(ouptut_dir):
    os.remove(os.path.join(ouptut_dir, f))


def _create_marksheet_func(roll_no, name, stud_res, corr_ans, pos_marks, neg_marks):
    wb = Workbook()
    sheet = wb.active
    # Sheet name
    sheet.title = 'quiz'
    # Adding image
    for c_name in list('ABCDE'):
        sheet.column_dimensions[c_name].width = 18
    img = Image('logo.png')
    sheet.add_image(img, 'A1')
    # Marks calc
    no_right_res = len(
        [res for i, res in enumerate(stud_res) if res == corr_ans[i]])
    no_wrng_res = len(
        [res for i, res in enumerate(stud_res) if res != corr_ans[i] and res != ''])
    no_na_res = len(['' for res in stud_res if '' == res])
    max_ques = len(corr_ans)
    total_marks_right_res = no_right_res * pos_marks
    total_marks_wrng_res = no_wrng_res * neg_marks
    total_marks = '{} / {}'.format(total_marks_right_res +
                                   total_marks_wrng_res, max_ques * pos_marks)
    # Adding title
    sheet.merge_cells('A6:E6')
    sheet.row_dimensions[6].height = 20
    title = sheet.cell(row=6, column=1)
    title_font_style = Font(name='Century', size='18',
                            underline='single',  bold=True)
    title.font = title_font_style
    title.value = 'Marksheet'
    title.alignment = Alignment(horizontal='center')
    # Font Styles and Border Styles
    basic_title_font_style = Font(name='Century', size='12', bold=True)
    basic_font_style = Font(name='Century', size='12')
    inp_corr_ans_font_style = Font(name='Century', size='12', color='018001')
    inp_wrng_ans_font_style = Font(name='Century', size='12', color='FF0D0D')
    corr_ans_font_style = Font(name='Century', size='12', color='0000FF')
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    # Student Details - Section
    sheet['A7'].value = 'Name:'
    sheet['A7'].alignment = Alignment(horizontal='right')
    sheet['A7'].font = basic_font_style
    sheet['B7'].value = name
    sheet['B7'].font = basic_title_font_style
    sheet['D7'].value = 'Exam:'
    sheet['D7'].alignment = Alignment(horizontal='right')
    sheet['D7'].font = basic_font_style
    sheet['E7'].value = sheet.title
    sheet['E7'].font = basic_title_font_style
    sheet['A8'].value = 'Roll Number:'
    sheet['A8'].alignment = Alignment(horizontal='right')
    sheet['A8'].font = basic_font_style
    sheet['B8'].value = roll_no
    sheet['B8'].font = basic_title_font_style
    # Marks - Section
    # Table - 1
    for row in sheet['A10:E13']:
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center')
    sheet['B10'].value = 'Right'
    sheet['B10'].font = basic_title_font_style
    sheet['C10'].value = 'Wrong'
    sheet['C10'].font = basic_title_font_style
    sheet['D10'].value = 'Not Attempt'
    sheet['D10'].font = basic_title_font_style
    sheet['E10'].value = 'Max'
    sheet['E10'].font = basic_title_font_style
    sheet['A11'].value = 'No.'
    sheet['A11'].font = basic_title_font_style
    sheet['A12'].value = 'Marking'
    sheet['A12'].font = basic_title_font_style
    sheet['A13'].value = 'Total'
    sheet['A13'].font = basic_title_font_style
    sheet['B11'].value = no_right_res
    sheet['B11'].font = inp_corr_ans_font_style
    sheet['B12'].value = pos_marks
    sheet['B12'].font = inp_corr_ans_font_style
    sheet['B13'].value = total_marks_right_res
    sheet['B13'].font = inp_corr_ans_font_style
    sheet['C11'].value = no_wrng_res
    sheet['C11'].font = inp_wrng_ans_font_style
    sheet['C12'].value = neg_marks
    sheet['C12'].font = inp_wrng_ans_font_style
    sheet['C13'].value = total_marks_wrng_res
    sheet['C13'].font = inp_wrng_ans_font_style
    sheet['D11'].value = no_na_res
    sheet['D11'].font = basic_font_style
    sheet['D12'].value = 0
    sheet['D12'].font = basic_font_style
    sheet['E11'].value = max_ques
    sheet['E11'].font = basic_font_style
    sheet['E13'].value = total_marks
    sheet['E13'].font = corr_ans_font_style
    # Response Table - Table - 2
    if max_ques > 25:
        sheet['A16'].value = 'Student Ans'
        sheet['A16'].font = basic_title_font_style
        sheet['A16'].border = border_style
        sheet['A16'].alignment = Alignment(horizontal='center')
        sheet['B16'].value = 'Correct Ans'
        sheet['B16'].font = basic_title_font_style
        sheet['B16'].border = border_style
        sheet['B16'].alignment = Alignment(horizontal='center')
        idx = 0
        for row in sheet['A17:B{}'.format(17+24)]:
            row[0].border = border_style
            row[0].alignment = Alignment(horizontal='center')
            row[1].border = border_style
            row[1].alignment = Alignment(horizontal='center')
            if stud_res[idx] == corr_ans[idx]:
                row[0].value = stud_res[idx]
                row[0].font = inp_corr_ans_font_style
            else:
                row[0].value = stud_res[idx]
                row[0].font = inp_wrng_ans_font_style
            row[1].value = corr_ans[idx]
            row[1].font = corr_ans_font_style
            idx += 1
        sheet['D16'].value = 'Student Ans'
        sheet['D16'].font = basic_title_font_style
        sheet['D16'].border = border_style
        sheet['D16'].alignment = Alignment(horizontal='center')
        sheet['E16'].value = 'Correct Ans'
        sheet['E16'].font = basic_title_font_style
        sheet['E16'].border = border_style
        sheet['E16'].alignment = Alignment(horizontal='center')
        for row in sheet['D17:E{}'.format(17+max_ques-26)]:
            row[0].border = border_style
            row[0].alignment = Alignment(horizontal='center')
            row[1].border = border_style
            row[1].alignment = Alignment(horizontal='center')
            if stud_res[idx] == corr_ans[idx]:
                row[0].value = stud_res[idx]
                row[0].font = inp_corr_ans_font_style
            else:
                row[0].value = stud_res[idx]
                row[0].font = inp_wrng_ans_font_style
            row[1].value = corr_ans[idx]
            row[1].font = corr_ans_font_style
            idx += 1
    else:
        sheet['A16'].value = 'Student Ans'
        sheet['A16'].font = basic_title_font_style
        sheet['A16'].border = border_style
        sheet['A16'].alignment = Alignment(horizontal='center')
        sheet['B16'].value = 'Correct Ans'
        sheet['B16'].font = basic_title_font_style
        sheet['B16'].border = border_style
        sheet['B16'].alignment = Alignment(horizontal='center')
        idx = 0
        for row in sheet['A17:B{}'.format(16+max_ques)]:
            row[0].border = border_style
            row[0].alignment = Alignment(horizontal='center')
            row[1].border = border_style
            row[1].alignment = Alignment(horizontal='center')
            if stud_res[idx] == corr_ans[idx]:
                row[0].value = stud_res[idx]
                row[0].font = inp_corr_ans_font_style
            else:
                row[0].value = stud_res[idx]
                row[0].font = inp_wrng_ans_font_style
            row[1].value = corr_ans[idx]
            row[1].font = corr_ans_font_style
            idx += 1
    # Save and create excel file
    wb.save(os.path.join(
        current_directory, 'assets', 'output', 'marksheet', '{}.xlsx'.format(roll_no.upper())))


def _create_concise_marksheet_func(res_data, corr_res,  pos_marks, neg_marks):
    with open(os.path.join(
            current_directory, 'assets', 'output', 'marksheet', 'concise_marksheet.csv'), 'w+', newline='') as f:
        writer = csv.writer(f)
        headers = ['Timestamp', 'Email address', 'Google_Score', 'Name', 'IITP webmail',
                   'Phone (10 digit only)', 'Score_After_Negative', 'Roll Number']
        headers.extend(['Unnamed']*len(corr_res))
        headers.append('statusAns')
        # Headers
        writer.writerow(headers)
        # Data - rows
        for res_row in res_data:
            data = []
            data.extend(res_row[0:2])
            # Marks calc
            no_right_res = len(
                [res for i, res in enumerate(res_row[7:]) if res == corr_res[i]])
            no_wrng_res = len(
                [res for i, res in enumerate(res_row[7:]) if res != corr_res[i] and res != ''])
            no_na_res = len(['' for res in res_row[7:] if '' == res])
            max_ques = len(corr_res)
            total_marks_right_res = no_right_res * pos_marks
            total_marks_wrng_res = no_wrng_res * neg_marks
            total_pos_marks = '{} / {}'.format(total_marks_right_res,
                                               max_ques * pos_marks)
            total_marks = '{} / {}'.format(total_marks_right_res +
                                           total_marks_wrng_res, max_ques * pos_marks)
            data.append(total_pos_marks)
            data.extend(res_row[3:6])
            data.append(total_marks)
            data.extend(res_row[6:])
            data.append([no_right_res, no_wrng_res, no_na_res])
            writer.writerow(data)


def marksheet_creator_run(ms_csv_filepath, r_csv_filepath, pos_marks, neg_marks, opt):
    try:
        with open(ms_csv_filepath, 'r') as f1, open(r_csv_filepath, 'r') as f2:
            master_roll_data = list(csv.reader(f1))[1:]
            responses_data = list(csv.reader(f2))[1:]
            all_roll_no = [el[0] for el in master_roll_data]
            for r in responses_data:
                marks = r[2].split(' / ')
                if marks[0] == marks[-1]:
                    corr_ans = responses_data[0][7:]
                if not r[6] in all_roll_no:
                    return 'No roll number with {} is present, Cannot Process!'.format(r[6]), 'red'
            if opt == 1:
                for row in master_roll_data:
                    r_data = list(
                        filter(lambda x: x[6] == row[0], responses_data))[0]
                    stud_res = r_data[7:]
                    if len(r_data):
                        _create_marksheet_func(
                            r_data[6], r_data[3], stud_res, corr_ans, pos_marks, neg_marks)
                    else:
                        _create_marksheet_func(
                            row[0], row[1], ['']*len(corr_ans), corr_ans, pos_marks, neg_marks)
                return 'Roll no. wise marksheet created successfully!', 'green'
            elif opt == 2:
                _create_concise_marksheet_func(
                    responses_data, corr_ans, pos_marks, neg_marks)
                temp_data = master_roll_data.copy()
                for row in responses_data:
                    temp_data.remove([row[6], row[3]])
                with open(os.path.join(
                        current_directory, 'assets', 'output', 'marksheet', 'concise_marksheet.csv'), 'a', newline='') as f:
                    writer = csv.writer(f)
                    for el in temp_data:
                        writer.writerow(['ABSENT', 'ABSENT', 'ABSENT', el[1], 'ABSENT',
                                        'ABSENT', 'ABSENT', el[0]])
                return 'Concise Marksheet created successfully!', 'green'

            else:
                return 'Something went wrong! Please try again.', 'red'

    except:
        return 'Something went wrong! Please try again.', 'red'

# Email attachment creator.


def send_email_marksheet(ms_csv_filepath, r_csv_filepath, pos_marks, neg_marks):
    marksheet_creator_run(ms_csv_filepath, r_csv_filepath,
                          pos_marks, neg_marks, 1)
    marksheet_creator_run(ms_csv_filepath, r_csv_filepath,
                          pos_marks, neg_marks, 2)
    return 'Emails sent successfully!', 'green'
